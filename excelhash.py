# created by wholl0p
# v1.0 30.09.2017
import hashlib
import openpyxl

md5 = hashlib.new("md5")
sha1 = hashlib.new("SHA1")

def create_hash_values(doc_path, DATA_COLUMN):
	wb = openpyxl.load_workbook(doc_path)

	sheet = wb.active

	for i in range(1, sheet.max_row + 1):
		cell_value = sheet.cell(row=i, column=DATA_COLUMN).value

		md5_hash_cell = sheet.cell(row=i, column=DATA_COLUMN + 1)
		sha1_hash_cell = sheet.cell(row=i, column=DATA_COLUMN + 2)

		md5.update(str(cell_value).encode())
		sha1.update(str(cell_value).encode())

		md5_hash_cell.value = md5.hexdigest()
		sha1_hash_cell.value = sha1.hexdigest()

		print(str(i) + ". " + md5.hexdigest() + " " + sha1.hexdigest())

	print(" ")
	print("1. neue Spalte: MD5, 2. neue Spalte: SHA1")
	wb.save(doc_path + "_updated.xlsx")

def main():
	print("!!! Die .xlsx Datei muss sich im selben Pfand befinden wie das Script!!!")
	doc_path = str(input("Bitte Name der Excel Datei eingeben (zb: Arbeitsmappe1.xlsx): "))
	DATA_COLUMN = int(input("Bitte Spalte der zu hashenden Werte eingeben (1,2,3,4,...): "))
	print(" ")

	if int(DATA_COLUMN) >= 1:
		create_hash_values(doc_path, DATA_COLUMN)

if __name__ == "__main__":
	main()
